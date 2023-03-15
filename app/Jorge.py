from fpdf import FPDF
import pandas as pd
import openpyxl



def data_cleaner(arg: any):
    if arg is None:
        return 'x'
    return arg



def data_access():

    print('Leyendo datos, por favor espere...')
    workbook = openpyxl.load_workbook('../fichas-Macro.xlsm')
    
    y=10

    # Selecciona la hoja de trabajo que contiene los datos
    sheet = workbook['DATOS-indice 2018']



    # Itera sobre cada fila y la imprime en la consola
    for index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True)):
        # Omitir las filas vacías
        if not any(row):
            continue
        
        if(index % 3 == 0):
            pdf.add_page()
            y=5; #reiniciamos el valor al por defecto en cada nueva página

        # pintar valores en el documento
        ticket_builder(
            num=data_cleaner(row[1],),
            names=data_cleaner(row[2]),
            last_names=data_cleaner(row[3]),
            father=data_cleaner(row[4]),
            mother=data_cleaner(row[5]),
            born_date=data_cleaner(row[6]),
            born_place=data_cleaner(row[7]),
            baptism_date=data_cleaner(row[8]),
            baptism_place=data_cleaner(row[9]),
            godfather=data_cleaner(row[10]),
            godmother=data_cleaner(row[11]),
            minister=data_cleaner(row[12]),
            y=y
        )

        # Después de la primera ficha incrementamos y sumamos el tamaño de la ficha y un margen de 5 mm
        y += 110
        # print('-------------------------------------------------------------------------------\n')

    # Cierra el archivo de Excel
    workbook.close()

def ticket_builder(num=0, names='', last_names='', born_date='', born_place='', baptism_date='', baptism_place='', godfather='', godmother='s', mother='', father='', y=10, minister='', hijo='Primero'):
    print('Contruyendo...')

    row_height = 4.8
    row_jump = 10

    # Cleaning data...

    complete_name = names.split(' ')
    complete_last_name = last_names.split(' ')

    first_name = 'x'
    second_name = 'x'

    first_last_name = 'x'
    second_last_name = 'x'

    first_name = complete_name[0]

    if len(complete_name) > 1:
        second_name = complete_name[1]

    first_last_name = complete_last_name[0]

    if len(complete_last_name) > 1:
        second_last_name = complete_last_name[1]

    complete_name_str = names + ' ' + last_names

    # region first block

    pdf.rect(x=5, y=y, w=200, h=110)

    pdf.cell(w=10, h=row_height, txt='No.', align='L',)
    pdf.set_font("Arial", size=11, style="I")
    pdf.cell(w=55, h=row_height, txt=str(num), align='C', border='B')
    pdf.set_font('Times', '', 10)

    # 2

    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    
    pdf.cell(w=30, h=row_height, txt='En la parroquia de:', align='L', border='0')
    
    pdf.cell(w=10, h=row_height, txt='', align='R', border='', )  
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=75, h=row_height, txt='San Isidro Labrador', align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=25, h=row_height, txt='1° Apellido', align='L',)
    pdf.set_font("Arial", size=11, style="I") 
    pdf.cell(w=40, h=row_height, txt=first_last_name, align='C', border='B')
    pdf.set_font('Times', '', 10)
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='C ', border='', )
    pdf.set_font("Arial", size=11, style="I") 
    pdf.cell(w=50, h=row_height, txt=baptism_place, align='C', border='B', )
    pdf.set_font('Times', '', 10)
    pdf.cell(w=10, h=row_height, txt='de', align='C', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=55, h=row_height, txt='Diócesis de Bluefields',
                   align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=25, h=row_height, txt='2° Apellido', align='L',)
    pdf.set_font("Arial", size=11, style="I") 
    pdf.cell(w=40, h=row_height, txt=second_last_name, align='C', border='B')
    pdf.set_font('Times', '', 10)
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='C', border='', )
    pdf.cell(w=30, h=row_height, txt='En la fecha de:', align='L',)
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=85, h=row_height, txt=baptism_date, align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=25, h=row_height, txt='1° Nombre', align='L',)
    pdf.set_font("Arial", size=11, style="I") 
    pdf.cell(w=40, h=row_height, txt=first_name, align='C', border='B')
    pdf.set_font('Times', '', 10)
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='C', border='', )
    pdf.cell(w=30, h=row_height, txt='El Ministro: ', align='L',)
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=85, h=row_height, txt=minister, align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=25, h=row_height, txt='2° Nombre', align='L',)
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=40, h=row_height, txt=second_name, align='C', border='B')
    pdf.set_font('Times', '', 10)
    pdf.cell(w=65, h=row_jump, txt='', align='C', border='0')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='C', border='')
    pdf.multi_cell(
        w=130, h=row_height, txt='Administró el Sacramento del Bautismo a:', align='C',)
    pdf.cell(w=75, h=row_jump, txt='', align='C', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=115, h=row_height, txt=complete_name_str, align='C', border='B')
    pdf.set_font('Times', '', 10)
    # endregion

    # region second block
    pdf.cell(
        w=65, h=row_height, txt='Recibió la primera comunión en la', align='L', border='0')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    pdf.cell(w=30, h=row_height, txt='Quién nació en: ', align='L', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=85, h=row_height, txt=born_place, align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=30, h=row_height, txt='parroquia de', align='L', border='0')
    pdf.cell(w=35, h=row_height, txt=' ', align='C', border='B')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    pdf.cell(w=30, h=row_height, txt='en la fecha de: ', align='L', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=85, h=row_height, txt=born_date, align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=10, h=row_height, txt='el', align='L', border='0')
    pdf.cell(w=25, h=row_height, txt='', align='L', border='B')
    pdf.cell(w=10, h=row_height, txt='de', align='L', border='0')
    pdf.cell(w=20, h=row_height, txt='', align='L', border='B')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    pdf.cell(w=10, h=row_height, txt='Hij: ', align='L', border='0')
    pdf.cell(w=5, h=row_height, txt="", align='C', border='B')
    pdf.cell(w=10, h=row_height, txt='de', align='C', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=85, h=row_height, txt=mother, align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=10, h=row_height, txt='del', align='L', border='0')
    pdf.cell(w=55, h=row_height, txt='', align='L', border='B')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    pdf.cell(w=10, h=row_height, txt='Y de: ', align='L', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=105, h=row_height, txt=father, align='L', border='B')
    pdf.set_font('Times', '', 10)
    pdf.multi_cell(w=20, h=row_jump, txt='', align='L', border='0')

    # endregion

    # region third block
    pdf.cell(w=65, h=row_height, txt='Contrajo matrimonio en la',
             align='L', border='T')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    pdf.cell(w=20, h=row_height, txt='Padrinos', align='L', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=95, h=row_height, txt=godfather, align='C', border='B')
    pdf.set_font('Times', '', 10)

    pdf.cell(w=30, h=5, txt='parroquia de', align='L', border='0')
    pdf.cell(w=35, h=row_height, txt=' ', align='C', border='B')
    # 2
    pdf.cell(w=10, h=row_height, txt='', align='L', border='0')
    pdf.cell(w=10, h=row_height, txt='Y', align='L', border='0')
    pdf.set_font("Arial", size=11, style="I") 
    pdf.multi_cell(w=105, h=row_height, txt=godmother, align='C', border='B')
    pdf.set_font('Times', '', 12)

    pdf.cell(w=15, h=row_height, txt='con', align='L', border='0')
    pdf.cell(w=50, h=row_height, txt='', align='L', border='B')
    # 2
    pdf.multi_cell(
        w=130, h=row_height, txt='A quienes se les advirtió su obligación y parentesco espiritual', border='0', align='C')
    pdf.set_font('Times', '', 10)
    pdf.cell(w=10, h=row_height, txt='el', align='L', border='0')
    pdf.cell(w=25, h=row_height, txt='', align='L', border='B')
    pdf.cell(w=10, h=row_height, txt='de', align='L', border='0')
    pdf.multi_cell(w=20, h=row_height, txt='', align='L', border='B')
    pdf.cell(w=10, h=row_height, txt='del', align='L', border='0')
    pdf.multi_cell(w=55, h=row_height, txt='', align='L', border='B')
    pdf.cell(w=10, h=row_height, txt='Nota: ', align='L', border='0')
    pdf.multi_cell(w=55, h=row_height, txt='', align='L', border='B')
    pdf.multi_cell(w=55, h=row_jump*2, txt='', align='L', border='0')

    # endregion

pdf = FPDF(orientation='P', unit='mm', format=(216, 360))

pdf.set_font('Times', '', 10)

data_access()

class PDF(FPDF):
    def footer(self):
        # Go to 1.5 cm from bottom
        self.set_y(-15)
        # Select Arial italic 8
        self.set_font('Arial', 'I', 8)
        # Print centered page number
        self.cell(0, 10, 'Page %s' % self.page_no(), 0, 0, 'C')

pdf.output('./report/fichas.pdf')
